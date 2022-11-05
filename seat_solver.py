from datetime import date, timedelta
import random
from pprint import pprint
import copy
from re import I, match
import time
from tqdm import tqdm
import csv
import pandas as pd
import openpyxl
import win32com.client


from Classes.Day import Day
from Classes.Teacher import Teacher
from Classes.Student import Student
from Classes.Subject import Subject
from Classes.Koma import Koma
from Classes.UniTable import UniTable
from Classes.MatchMatrix import MatchMatrix
from Classes.Error import *
from utils import *


def register_period():
    """
    講習期間を登録する。
    """
    start_day = date(2021, 7, 1)
    end_day = date(2021, 8, 31)
    return start_day, end_day


def register_teachers():
    """
    講師を登録する。
    teachers = {'00001' : Teacher
                '00002' : Teacher} という形式の dict
    """
    teachers = {}
    with open('data/teachers/info.csv', encoding='utf-8') as f:
        f.readline() # 最初の行は読み飛ばす。
        for line in f:
            rows = line.rstrip('\n').split(',')
            id, name = rows
            teacher = Teacher(name=name, id=id)
            schedule_path = f'data/teachers/schedules/{id}.csv'

            # スケジュール関係の登録。
            teacher.set_schedule_from_csv(schedule_path)
            teacher.set_schedule_ava()

            # UniTable.unitable のもととなる table の作成。
            teacher.make_table()

            teachers[id] = teacher
    return teachers


def register_subjects(student, subjects_path):
    """
    register_students() の中で呼ばれる。
    subjects の情報を1行ずつ見て、student に登録する。
    """ 
    with open(subjects_path, encoding='utf-8')as s:
        s.readline()
        for subject_info in s:
            subject_info = subject_info.rstrip('\n').split(',')
            subject, num_of_lecture, is_one_on_one = subject_info
            subject_ = Subject(student.name, student.id, subject, int(num_of_lecture), int(is_one_on_one))
            student.register_subject(subject_)
        return student


def register_students():
    """
    生徒を登録する。
    また、教科を担当する講師に紐づける。
    students = {'001' : Student
                '002' : Student} という形式の dict
    """
    students = {}
    with open('data/students/info.csv', encoding='utf-8') as f:
        f.readline() # 最初の行は読み飛ばす。
        for student_info in f:
            student_info = student_info.rstrip('\n').split(',')
            id, name = student_info
            student = Student(name=name, id=id)

            # このStudentインスタンスの情報に関係するパスを生成する。
            schedule_path = f'data/students/schedules/{id}.csv'
            subjects_path = f'data/students/subjects/{id}.csv'
            
            # スケジュール関係の登録。
            student.set_schedule_from_csv(schedule_path)
            student.set_schedule_ava()

            # 教科情報を生徒に登録。
            student = register_subjects(student, subjects_path)

            # Subject が student にすべて登録されたら、1授業ずつに分解し、lectures リストに追加。
            student.set_lectures()
            
            students[id] = student
    return students


def arrange_match_matrix(teachers, students):
    """
    match_matrix を作る。
    その後、既存の担当を登録する。
    """
    match_matrix = MatchMatrix(teachers, students)
    match_matrix.register_existing_charge()
    return match_matrix

def uni_register(uni_table):
    """
    授業がcancelされたSeatTableに対して、再度授業を登録する。
    登録後、SeatTableはスコアを計算して返される。
    """
    
    for jigen in uni_table.ava_komas: 
        ava_stu = uni_table.ava_stu_dic[jigen]
        if len(ava_stu) == 0:
            continue
        ava_tea = uni_table.get_ava_tea(jigen)
        if len(ava_tea) == 0:
            continue
        slcted_tea = random.choice(ava_tea)
        state = uni_table.unitable[jigen][slcted_tea].state
        if state == 1:
            # ------------------------
            # 授業2の登録ここから
            try:
                uni_table, slcted_stu, _ = uni_choice_and_register_lec(jigen, uni_table, slcted_tea, ava_stu) # 登録
            except CannotRegisterError:
                continue
            if uni_table.is_all_lec_rgstred(slcted_stu): # ある生徒の授業が全登録されたか判定
                uni_table.remove_stu(slcted_stu)
            if uni_table.is_able_to_finish(): # 全生徒の授業が全登録されたか判定
                break
            # 授業2の登録ここまで
            # ------------------------
        elif state == 0:
            # ------------------------
            # 授業1の登録ここから
            uni_table, slcted_stu, ava_stu = uni_choice_and_register_lec(jigen, uni_table, slcted_tea, ava_stu) # 登録
            if uni_table.is_all_lec_rgstred(slcted_stu): # ある生徒の授業が全登録されたか判定
                uni_table.remove_stu(slcted_stu)
            if uni_table.is_able_to_finish(): # 全生徒の授業が全登録されたか判定
                break
            # 授業1の登録ここまで
            # ------------------------

            if uni_table.is_katakoma(slcted_tea, jigen) and is_there_anyone(ava_stu): # (片コマ) and (受講可能な生徒が他にもいる)
                # ------------------------
                # 授業2の登録ここから
                try:
                    uni_table, slcted_stu, _ = uni_choice_and_register_lec(jigen, uni_table, slcted_tea, ava_stu) # 登録
                except CannotRegisterError:
                    continue
                if uni_table.is_all_lec_rgstred(slcted_stu): # ある生徒の授業が全登録されたか判定
                    uni_table.remove_stu(slcted_stu)
                if uni_table.is_able_to_finish(): # 全生徒の授業が全登録されたか判定
                    break
                # 授業2の登録ここまで
                # ------------------------
    uni_table.calculate_score()
    return uni_table

def uni_choice_and_register_lec(jigen, uni_table, tea_id, ava_stu):
    """
    授業の登録処理を行う。
    具体的には、次のような操作を行う。
    
    -----------
    1. 生徒をランダム選択
    2. 授業をランダム選択
    3. 選んだ授業を SeatTable.table 内の jigen で指定された日時 に登録
    4. 登録した授業を SeatTable.left_lecs 内の 選択された生徒と結びついた授業 から削除
    5. 登録した生徒を ava_stu から削除
    """
    slcted_stu = random.choice(ava_stu) # 生徒ID がrandom.choice される
    slcted_lec = random.choice(uni_table.left_lecs[slcted_stu]) # Lecture インスタンスがrandom.choice される
    try:
        uni_table.unitable[jigen][tea_id].register(slcted_lec) # Komaオブジェクトのクラスメソッドregisterの呼び出し
        uni_table.left_lecs[slcted_stu].remove(slcted_lec) # 登録したLectureをseat_table.left_lecs から削除
        ava_stu.remove(slcted_stu) # 登録した生徒をava_stu_から削除(1人の生徒が対2コマの両方に登録されるのを防ぐ)
        return uni_table, slcted_stu, ava_stu
    except CannotRegisterError:
        raise # CannotRegisterError をそのまま呼び出し元に送信


def uni_first_search(unitable, n_search=1000):
    """
    一回目の探索を行う。
    """
    uni_tables = []
    for i in tqdm(range(n_search)):
        uni_table = copy.deepcopy(unitable)
        uni_table.shuffle_ava_komas()
        uni_table = uni_register(uni_table)
        uni_tables.append(uni_table)
    return uni_tables

@profile
def uni_step_generation(tables, students, n_cancel=2, n_resolve=10, n_keeptable=50):
    """
    世代を1つ進める。
    """
    # tableをスコア順に並べ替えて、上位 n_keeptable 個に厳選
    good_tables = squeeze_tables(tables, n_keeptable)

    resolved_tables = []
    for uni_table in tqdm(good_tables):
        
        # score をリセット
        uni_table.reset_score()
        
        # cancel せずに resolve
        resolved_tables.append(uni_register(copy.deepcopy(uni_table)))

        # cancel してから resolve
        for i in range(n_cancel): # n_cancel 通りキャンセル
            uni_table_i = copy.deepcopy(uni_table)
            uni_table_i.cancel_lecs(students)
            for j in range(n_resolve): # n_resolve 通り再登録
                uni_table_j = copy.deepcopy(uni_table_i)
                uni_table_j = uni_register(uni_table_j)
                resolved_tables.append(uni_table_j)
    return resolved_tables


def uni_process_generation(uni_table, students, N_SEARCH, N_GENERATIONS, N_CANCEL, N_RESOLVE, N_KEEPTABLE):
    """
    UniTable に対応している遺伝的処理。
    """
    # 1st search
    print(f"\n\t===1st generation===")
    
    table_g = uni_first_search(uni_table, N_SEARCH)
    best_table = squeeze_tables(table_g, 1)
    print_unitable(best_table)

    # Search loop after 1st search    
    for search_counter in range(N_GENERATIONS-1):
        print(f"\n\t==={search_counter+2}-th generation===")
        table_g = uni_step_generation(table_g, students, N_CANCEL, N_RESOLVE, N_KEEPTABLE)
        best_table = squeeze_tables(table_g, 1)
        print_unitable(best_table)
    return best_table

def devide_unitable_by_day(unitable):
    """
    input: unitable
    output: list of dict
    """
    li = []
    while True:
        day = list(unitable.unitable.keys())[0][0:10] # リストのうち最初のjigenの'[YYYY-MM-DD]' を抽出
        same_days = {jigen: komas for jigen, komas in unitable.unitable.items() if day in jigen} # day が key に入っていれば、同じ日と判定
        li.append(same_days)
        for k in same_days.keys():
            del unitable.unitable[k]
        if len(unitable.unitable) == 0:
            break
    return li


def demo(unitable, students):
    N_SEARCH, N_CANCEL, N_RESOLVE, N_KEEPTABLE, N_GENERATIONS = decide_constants(development_flag=1)
    best_table = uni_process_generation(unitable, students, N_SEARCH, 10, N_CANCEL, N_RESOLVE, N_KEEPTABLE)
    li = devide_unitable_by_day(best_table)
    
    wb = openpyxl.load_workbook("demo/template.xlsx")
    tps = wb['Sheet1'] # テンプレート

    for i, jigens in enumerate(li):
        ws = wb.copy_worksheet(wb['Sheet1'])
        ws.title = list(jigens.keys())[0][0:10]
        c_title = ws["A1"]
        day= list(jigens.keys())[0][0:10]
        c_title.value = f'{day}の座席表'
        r_jigen = 3 # 3列目を基準
        
        for jigen, komas in jigens.items():
            c1 = ws.cell(row=r_jigen, column=2)
            c1._style = tps.cell(row=3, column=2)._style
            c1.value = f"{jigen[-1]}限"
            c2 = ws.cell(row=r_jigen, column=3)
            c2._style = tps.cell(row=3, column=3)._style
            c2.value = "講師"
            c3 = ws.cell(row=r_jigen, column=4)
            c3._style = tps.cell(row=3, column=4)._style
            c3.value = "生徒1"
            c4 = ws.cell(row=r_jigen, column=5)
            c4._style = tps.cell(row=3, column=5)._style
            c4.value = "生徒2"
            for i, (tea, koma) in enumerate(komas.items()):
                ct = ws.cell(row=r_jigen+i+1, column=3)
                ct.value = f"講師{tea}"
                ct._style = tps.cell(row=4, column=3)._style
                ck1 = ws.cell(row=r_jigen+i+1, column=4)
                ck1._style = tps.cell(row=4, column=4)._style
                ck1.value = "" if koma.lecture_1 == None else str(koma.lecture_1) # None にならないことを意図しているが None になる
                ck2 = ws.cell(row=r_jigen+i+1, column=5)
                ck2._style = tps.cell(row=4, column=5)._style
                ck2.value = "" if koma.lecture_2 == None else str(koma.lecture_2)
            r_jigen += 4

    wb.remove(wb['Sheet1'])
    wb.save("demo/output.xlsx")
    
    # xlsx ---> pdf への変換
    excel = win32com.client.Dispatch("Excel.Application") # Excelを起動
    path_exc = r"C:\Users\iojdw\GitHub\SeatTableSolver\demo\output.xlsx"
    path_pdf = r"C:\Users\iojdw\GitHub\SeatTableSolver\demo\output.pdf"
    book = excel.Workbooks.Open(path_exc) # Excelでファイルを読み込み
    book.ExportAsFixedFormat(0, path_pdf) # PDF形式で保存
    excel.Quit() # Excelを終了
    

                
def main():
    start_time = time.time()
    
    # 日付情報, 講師情報, 生徒情報の登録
    start_day, end_day = register_period()
    teachers = register_teachers()
    students = register_students()
    
    # UniTable の初期作成
    unitable = UniTable(teachers, students)
    
    # demo(unitable, students)
    # 評価用情報を pickle 形式で保存
    dump_combi_spec()
    dump_existing_charge()
    match_matrix = arrange_match_matrix(teachers, students)
    
    # 定数の決定
    N_SEARCH, N_CANCEL, N_RESOLVE, N_KEEPTABLE, N_GENERATIONS = decide_constants(development_flag=1) # SeatTable の cancel_rate も定数

    
    # unitable に対して、uni_first_search と uni_step_generation を行う
    
    best_table = uni_process_generation(unitable, students, N_SEARCH, N_GENERATIONS, N_CANCEL, N_RESOLVE, N_KEEPTABLE)
        
    elapsed_time = time.time() - start_time
    print(f'\n探索時間: {elapsed_time:.3f} 秒')
    


if __name__ == '__main__':
    main()
